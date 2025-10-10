from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment

from excel.core.dispatcher import Dispatcher
from excel.core.models.parse_result import CellparseResult, CI00ReadParseResult, PL10ReadParseResult
from excel.handlers.writer.write_handler_base import WriteHandlerBase

@Dispatcher.register_handlers
class TotalHandlers(WriteHandlerBase):
    """
    合计数处理器
    """

    @Dispatcher.keyword("WRITE_TOTAL GROSS WEIGHT:")
    def handle_total_gross_weight(self, cell: Cell) -> CellparseResult:
        """
        处理[TOTAL_GROSS_WEIGHT,TOTAL_NET_WEIGHT,TOTAL_CARTONS,TOTAL_MEASUREMENT]显示
        """
        self._worksheet.cell(cell.row, cell.column + 1, self._datasource.get_data_source(PL10ReadParseResult)["total_gross_weight"])
        self._worksheet.cell(cell.row + 1, cell.column + 1, self._datasource.get_data_source(PL10ReadParseResult)["total_net_weight"])
        self._worksheet.cell(cell.row + 2, cell.column + 1, self._datasource.get_data_source(PL10ReadParseResult)["total_packages"])
        self._worksheet.cell(cell.row + 3, cell.column + 1, self._datasource.get_data_source(PL10ReadParseResult)["total_measurement"])
        return CellparseResult(next_row_index=cell.row + 4)

    @Dispatcher.keyword("WRITE_TRADE TERMS:CIP EGYPT")
    def handle_total_amount_usd_english(self, cell: Cell) -> CellparseResult:
        """
        处理美元合计英文显示(货代Invoice和With material code Sheet用)
        """
        for sheetname in ["货代 Invoice", "With material code"]:
            self._workbook[sheetname].merge_cells(start_row=cell.row + 2, start_column=cell.column, end_row=cell.row + 2, end_column=cell.column + 4)
            self._workbook[sheetname].cell(cell.row + 2, cell.column).alignment = Alignment(wrap_text=True, vertical="center")
            self._workbook[sheetname].row_dimensions[cell.row + 2].height = 30
            self._workbook[sheetname].cell(cell.row + 2, cell.column).value = self._datasource.get_data_source(CI00ReadParseResult)["total_amount_english"]

        return CellparseResult()
