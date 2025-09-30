from typing import List

from injector import inject
from openpyxl.cell.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from excel.core.dispatcher import Dispatcher
from excel.core.models.parse_result import CellparseResult, CI00ReadParseResult
from excel.core.models.purchase_detail import CI00PurchaseDetail
from excel.handlers.models.pending_file_model import PendingFileModel
from excel.handlers.models.writer_datasource import WriterDataSource
from excel.handlers.services.authenticationed_phone_model_service import AuthenticationedPhonemodelService
from excel.handlers.services.battery_brand_service import BatteryBrandService
from excel.handlers.services.hscode_service import HScodeService
from excel.handlers.writer.purchase_detail_handlers import WritePurchasedetailHandlers

@Dispatcher.register_handlers
class WritePurchasedetailInvoiceHandlers(WritePurchasedetailHandlers):
    """
    [货代 Invoice]采购明细处理器
    """

    @inject
    def __init__(
            self,
            workbook: Workbook,
            worksheet: Worksheet,
            datasource: WriterDataSource,
            pending_file_model: PendingFileModel,
            hscode_service: HScodeService,
            battery_brand_service: BatteryBrandService,
            authentication_phonemodel_service: AuthenticationedPhonemodelService):
        """
        初始化E[货代 Invoice]采购明细处理器
        :param workbook: 要写入的Workbook
        :param worksheet: 要写入的Worksheet
        :param datasource: Excel写入器数据源
        :param pending_file_model: 待处理文件模型
        """
        super().__init__(workbook, worksheet, datasource)
        self._pending_file_model = pending_file_model
        self._hscode_service = hscode_service
        self._battery_brand_service = battery_brand_service
        self._authentication_phonemodel_service = authentication_phonemodel_service

    @Dispatcher.keyword("WRITE_HS")
    def handle_invoice_purchase_detail(self, cell: Cell) -> CellparseResult:
        """
        处理[货代Invoice] Sheet采购明细
        """
        purchase_details = self._datasource.get_data_source(CI00ReadParseResult)["purchase_details"]
        self._insert_blank_rows(self._worksheet, cell.row + 2, len(purchase_details))
        self._insert_blank_rows(self._workbook["With material code"], cell.row + 2, len(purchase_details))

        # 插入采购明细
        self._insert_purchase_details(self._worksheet, cell, purchase_details)
        self._insert_purchase_details(self._workbook["With material code"], cell, purchase_details)

        return CellparseResult(next_row_index=cell.row + 2 + len(purchase_details))

    def _insert_purchase_details(self, worksheet: Worksheet, cell: Cell, purchase_details: List[CI00PurchaseDetail]):
        """
        插入电池 [货代Invoice] 采购明细
        [hs_code, model, description of goods, quantity, unit price, amount usd, (oppo电池:brand, oppo非电池:origin_country, rmg:remark)]
        """
        # 已认证手机型号工厂编码从头开始
        self._authentication_phonemodel_service.to_first(self._pending_file_model.brand_category)

        # 循环插入所有采购明细
        for index, purchase_detail in enumerate(purchase_details):
            worksheet.cell(cell.row + 2 + index, 1).value = self._hscode_service.get_hscode(purchase_detail["material_code"])
            worksheet.cell(cell.row + 2 + index, 3).value = purchase_detail["description"]
            worksheet.cell(cell.row + 2 + index, 5).value = purchase_detail["quantity"]
            worksheet.cell(cell.row + 2 + index, 6).value = purchase_detail["unit_price"]
            worksheet.cell(cell.row + 2 + index, 7).value = purchase_detail["amount_usd"]

            # Model列、第八列[brand/remark/country of origin]
            if self._pending_file_model.brand_subcategory == "电池":
                worksheet.cell(cell.row + 2 + index, 2).value = self._battery_brand_service.get_battry_brand(purchase_detail["material_code"]).model
                worksheet.cell(cell.row + 2 + index, 8).value = self._battery_brand_service.get_battry_brand(purchase_detail["material_code"]).brand
            elif self._pending_file_model.factory_name == "RMG":
                worksheet.cell(cell.row + 2 + index, 2).value = self._authentication_phonemodel_service.get_next_factorycode(self._pending_file_model.brand_category)
                worksheet.cell(cell.row + 2 + index, 8).value = purchase_detail["remark"]
            else:
                worksheet.cell(cell.row + 2 + index, 2).value = self._authentication_phonemodel_service.get_next_factorycode(self._pending_file_model.brand_category)
                worksheet.cell(cell.row + 2 + index, 8).value = purchase_detail["origin_country"]

            # [With material code] Sheedt [Material Code] 列
            worksheet.cell(cell.row + 2 + index, 9).value = purchase_detail["material_code"]
