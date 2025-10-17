from typing import List, Dict

from injector import inject
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from excel.core.dispatcher import Dispatcher
from excel.core.models.parse_result import PL10ReadParseResult, CellparseResult
from excel.handlers.models.pending_file_model import PendingFileModel
from excel.handlers.models.writer_datasource import WriterDataSource
from excel.handlers.services.authenticationed_phone_model_service import AuthenticationedPhonemodelService
from excel.handlers.writer.purchase_detail_handlers import WritePurchasedetailHandlers
from excel.handlers.writer.write_handler_base import WriteHandlerBase

@Dispatcher.register_handlers
class WritePurchasedetailPackingHandlers(WritePurchasedetailHandlers, WriteHandlerBase):
    """
    [货代 Packing]采购明细处理器
    """

    @inject
    def __init__(
            self,
            workbook: Workbook,
            worksheet: Worksheet,
            datasource: WriterDataSource,
            pending_file_model: PendingFileModel,
            authentication_phonemodel_service: AuthenticationedPhonemodelService):
        """
        初始化E[货代 Invoice]采购明细处理器
        :param workbook: 要写入的Workbook
        :param worksheet: 要写入的Worksheet
        :param datasource: Excel写入器数据源
        :param pending_file_model: 待处理文件模型
        :param authentication_phonemodel_service: 已认证手机型号服务
        """
        super().__init__(workbook, worksheet, datasource)
        self._pending_file_model = pending_file_model
        self._authentication_phonemodel_service = authentication_phonemodel_service

    @Dispatcher.keyword("WRITE_SHIPPING MARKS")
    def handle_packing_purchase_detail(self, cell: Cell) -> CellparseResult:
        """
        处理[货代Packing] Sheet采购明细
        """
        pl10_read_parse_result = self._datasource.get_data_source(PL10ReadParseResult)
        self._insert_blank_rows(self._worksheet, cell.row + 1, len(pl10_read_parse_result["purchase_details"]))
        self._insert_purchase_details(cell, pl10_read_parse_result)

        # 写Total值
        self._write_total(self._worksheet.cell(cell.row + 2 + len(pl10_read_parse_result["purchase_details"]), 3))

        return CellparseResult(next_row_index=self._worksheet.max_row + 1)

    def _insert_purchase_details(self, cell: Cell, pl10_read_parse_result: PL10ReadParseResult):
        """
        插入 [货代 Packing] 采购明细
        [Shipping Makrs, Model, Description of goods,PLTS(Oppo工厂有), PCS/CTN, CTNS, Total Quantity, N.W(KG), G.W(KG)]
        """
        # 已认证手机型号工厂编码从头开始
        self._authentication_phonemodel_service.to_first(self._pending_file_model.brand_category)

        # 循环插入所有采购明细
        merged_total_packages: List[Dict] = []
        total_column_start = 5 if self._pending_file_model.factory_name == "OPPO" else 4
        for index, purchase_detail in enumerate(pl10_read_parse_result["purchase_details"], start=0):
            self._worksheet.cell(cell.row + 1 + index, 1).value = purchase_detail["shipping_marks"]
            self._worksheet.cell(cell.row + 1 + index, 2).value = self._authentication_phonemodel_service.get_next_factorycode(self._pending_file_model.brand_category)
            self._worksheet.cell(cell.row + 1 + index, 3).value = purchase_detail["description"]
            self._worksheet.cell(cell.row + 1 + index, column=total_column_start).value = purchase_detail["total_quantity"]
            self._worksheet.cell(cell.row + 1 + index, column=total_column_start + 2).value = purchase_detail["total_quantity"]

            # 计算CTNS列值合并
            if purchase_detail["total_packages"] is not None:
                merged_total_packages.append({
                    "start_cell": self._worksheet.cell(row=cell.row + 1 + index, column=total_column_start + 1),
                    "count": 1,
                    "value": purchase_detail["total_packages"]})
            else:
                merged_total_packages[-1]["count"] += 1
            #
            self._worksheet.cell(cell.row + 1 + index, column=total_column_start + 1).value = purchase_detail["total_packages"]

        # 写N.W(KG), G.W(KG), CTNS
        self._write_nw_and_gw(cell, total_column_start, pl10_read_parse_result)
        self._write_ctns_with_merge(merged_total_packages)

    def _write_nw_and_gw(self, cell: Cell, total_column_start: int, pl10_read_parse_result: PL10ReadParseResult):
        """
        写N.W(KG)和G.W(KG)
        """
        purchase_details_count = len(pl10_read_parse_result["purchase_details"])
        self._worksheet.merge_cells(start_row=cell.row + 1, start_column=total_column_start + 3, end_row=cell.row + purchase_details_count, end_column=total_column_start + 3)
        self._worksheet.merge_cells(start_row=cell.row + 1, start_column=total_column_start + 4, end_row=cell.row + purchase_details_count, end_column=total_column_start + 4)
        self._worksheet.cell(cell.row + 1, column=total_column_start + 3).alignment = Alignment(vertical="center", horizontal="center")
        self._worksheet.cell(cell.row + 1, column=total_column_start + 3).value = pl10_read_parse_result["total_net_weight"]
        self._worksheet.cell(cell.row + 1, column=total_column_start + 4).alignment = Alignment(vertical="center", horizontal="center")
        self._worksheet.cell(cell.row + 1, column=total_column_start + 4).value = pl10_read_parse_result["total_gross_weight"]

    def _write_ctns_with_merge(self, merged_total_packages: List[Dict]):
        """
        写入合并后的ctns值
        """
        for item in merged_total_packages:
            if item["count"] > 1:
                self._worksheet.merge_cells(
                    start_row=item["start_cell"].row,
                    start_column=item["start_cell"].column,
                    end_row=item["start_cell"].row + item["count"] - 1,
                    end_column=item["start_cell"].column)
            item["start_cell"].value = item["value"]

    def _write_total(self, cell: Cell):
        """
        写Total值
        """
        total_column_start = 3 if self._pending_file_model.factory_name == "OPPO" else 2
        self._worksheet.cell(cell.row, cell.column + total_column_start, self._datasource.get_data_source(PL10ReadParseResult)["total_packages"])
        self._worksheet.cell(cell.row, cell.column + total_column_start + 1, self._datasource.get_data_source(PL10ReadParseResult)["total_quantity"])
        self._worksheet.cell(cell.row, cell.column + total_column_start + 2, self._datasource.get_data_source(PL10ReadParseResult)["total_net_weight"])
        self._worksheet.cell(cell.row, cell.column + total_column_start + 3, self._datasource.get_data_source(PL10ReadParseResult)["total_gross_weight"])
