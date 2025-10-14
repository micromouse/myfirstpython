import re
from typing import List

from injector import inject
from openpyxl.worksheet.worksheet import Worksheet

from excel.appsettings import AppSettings
from excel.core.fluent_excel_writer import FluentExcelWriter
from excel.core.utils import Utils
from excel.handlers.models.pending_file_model import PendingFileModel

class RegisteredInvoicNumberService:
    """
    已注册发票号服务
    """
    _registered_invoice_numbers: List[str] = []

    @inject
    def __init__(self, appsettings: AppSettings):
        """
        初始化已注册发票号服务
        :param appsettings: 应用程序配置
        """
        self._appsettings = appsettings
        if not self._registered_invoice_numbers:
            RegisteredInvoicNumberService._registered_invoice_numbers = \
                Utils.get_excel_column_values(self._appsettings.registered_invoice_number_file, 0, 3, 2)

        # 插入一个空行分隔
        if len(self._registered_invoice_numbers) > 0:
            FluentExcelWriter() \
                .set_file(self._appsettings.registered_invoice_number_file) \
                .write(lambda sheet: sheet.cell(sheet.max_row + 1, 1, ""))

    def save_new_invoice_number(
            self,
            pending_file_model: PendingFileModel,
            original_invoice_number: str,
            new_invoice_number: str):
        """
        保存新发票号到已注册发票号文件
        :param pending_file_model: 待处理采购CI&PL文件模型
        :param original_invoice_number: 新发票号
        :param new_invoice_number: 新发票号
        """

        def _add_new_invoice_row(sheet: Worksheet):
            """
            添加新发票号行
            :param sheet: 已注册发票号Excel Sheet
            """
            new_row_index = sheet.max_row + 1
            sheet.cell(new_row_index, 1, pending_file_model.pending_file_path.name[:4])
            sheet.cell(new_row_index, 2, pending_file_model.brand_subcategory)
            sheet.cell(new_row_index, 3, new_invoice_number)
            sheet.cell(new_row_index, 4, original_invoice_number)

        FluentExcelWriter() \
            .set_file(self._appsettings.registered_invoice_number_file) \
            .write(_add_new_invoice_row)

    def get_new_invoice_number(self, current_invoice_nubmers: List[str]) -> str:
        """
        获得新发票号
        :param current_invoice_nubmers: 当前发票号集合
        :return: 新发票号
        """
        unused_invoice_numbers = sorted((set(current_invoice_nubmers) - set(self._registered_invoice_numbers)), reverse=True)

        # 有未使用的发票号，取最大的一个
        if unused_invoice_numbers:
            new_invoice_number = unused_invoice_numbers[0]
        else:
            new_invoice_number = current_invoice_nubmers[0].strip()
            used_invoice_numbers = list({
                n for n in self._registered_invoice_numbers if
                n == new_invoice_number or
                n.startswith(f"{new_invoice_number}-")
            })
            new_invoice_number = f"{new_invoice_number}-{self._get_max_suffix_number(new_invoice_number, used_invoice_numbers) + 1}"

        return new_invoice_number

    @staticmethod
    def _get_max_suffix_number(base_invoice_number: str, used_invoice_numbers: List[str]) -> int:
        """
        从已使用发票号集合中获得最大后缀号
        :param base_invoice_number: 基础发票号
        :param used_invoice_numbers: 已使用发票号
        :return: 已使用发票号最大后缀
        """
        # re.escape(base_invoice_number)会把[base_invoice_number]里的特殊字符自动转义。如:
        # re.escape("a-b-c.006")结果是"a\\-b\\-c\\.006"
        pattern = re.compile(rf'^{re.escape(base_invoice_number)}-(\d+)$')

        """
        re.match(pattern, string) → 从字符串开头匹配，只匹配开头
        re.search(pattern, string) → 整个字符串搜索，匹配任意位置
        re.fullmatch(pattern, string) → 整个字符串完全匹配
        这行代码用了 Python 3.8+ 的“海象运算符”（walrus operator :=），可以顺便在条件判断里给变量赋值，然后在生成器里直接使用。
        suffix_numbers = []
        for n in used_invoice_numbers:
            m = pattern.fullmatch(n)  # 匹配 base-数字
            if m:                     # 如果匹配成功
                suffix_numbers.append(int(m.group(1)))  # 提取尾号数字                
        """
        suffix_numbers: List[int] = [int(m.group(1)) for n in used_invoice_numbers if (m := pattern.fullmatch(n))]
        return max(suffix_numbers, default=0)
