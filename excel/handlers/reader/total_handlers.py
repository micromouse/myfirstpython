from decimal import Decimal

from openpyxl.cell.cell import Cell

from excel.core.dispatcher import Dispatcher
from excel.core.models.parse_result import CellparseResult
from excel.core.utils import Utils
from excel.handlers.reader.read_handler_base import ReadhandleBase

@Dispatcher.register_handlers
class ReadTotalHandlers(ReadhandleBase):
    """
    合计数处理器
    """

    @Dispatcher.keyword("READ_TOTAL :")
    def handle_total(self, cell: Cell) -> CellparseResult:
        """
        处理合计值
        """
        total_quantity = int(float(str(self._worksheet.cell(cell.row, cell.column + 1).value).replace(",", "")))
        total_amount = Decimal(str(self._worksheet.cell(cell.row, cell.column + 3).value).replace(",", ""))
        return CellparseResult({
            "total_quantity": total_quantity,
            "total_amount": total_amount,
            "total_amount_english": Utils.format_currency_amount(total_amount)
        })

    @Dispatcher.keyword("READ_TOTAL GROSS WEIGHT")
    def handle_total_gross_weight(self, cell: Cell) -> CellparseResult:
        """
        处理 [TOTAL GROSS WEIGHT][TOTAL NET WEIGHT][TOTAL PACKAGES][TOTAL QUANTITY][TOTAL MEASUREMENT]
        """
        total_gross_weight = Decimal(str(self._worksheet.cell(cell.row, cell.column + 6).value).replace(",", ""))
        total_net_weight = Decimal(str(self._worksheet.cell(cell.row + 1, cell.column + 6).value).replace(",", ""))
        total_packages = int(float(str(self._worksheet.cell(cell.row + 2, cell.column + 6).value).replace(",", "")))
        total_quantity = int(float(str(self._worksheet.cell(cell.row + 3, cell.column + 6).value).replace(",", "")))
        total_measurement = Decimal(str(self._worksheet.cell(cell.row + 4, cell.column + 6).value).replace(",", ""))
        return CellparseResult({
            "total_gross_weight": total_gross_weight,
            "total_net_weight": total_net_weight,
            "total_packages": total_packages,
            "total_quantity": total_quantity,
            "total_measurement": total_measurement
        }, next_row_index=self._worksheet.max_row + 1)
