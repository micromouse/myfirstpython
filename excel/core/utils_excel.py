from contextlib import contextmanager
from pathlib import Path
from typing import List, Callable, Tuple, Optional, Union

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from excel.core.models.write_excel_model import WriteExcelModel

class UtilsExcel:
    """
    excel工具 - excel
    """

    @staticmethod
    @contextmanager
    def open_workbook(file: Union[str, Path], **kwargs):
        """
        支持资源释放的打开Excel Workbook
        :param file: 文件名
        :param kwargs: 参数
        :return: Excel Workbook
        """
        workbook = load_workbook(file, **kwargs)
        try:
            yield workbook
        finally:
            workbook.close()

    @staticmethod
    def load_worksheet(
            file: str,
            sheet_name: Optional[str] = None,
            sheet_index: Optional[int] = None,
            **kwargs
    ) -> Tuple[Workbook, Worksheet]:
        """
        载入Excel Worksheet
        :param file: Excel文件名
        :param sheet_name: (可选)Worksheet名
        :param sheet_index: (可选)Worksheet索引
        :param kwargs: 参数集合
        :return: [Workbok, Worksheet]元组
        """
        if not Path.exists(Path(file)):
            raise FileNotFoundError(f"文件[{file}]不存在")

        workbook = load_workbook(file, **kwargs)

        # 传递了Sheet名
        if sheet_name:
            if sheet_name not in [sheet.title for sheet in workbook.worksheets]:
                raise KeyError(f"文件[{file}]中不存在 [{sheet_name}] Excel Worksheet")
            return workbook, workbook[sheet_name]

        # 传递了Sheet索引
        if sheet_index > -1:
            if sheet_index < 0 or sheet_index >= len(workbook.worksheets):
                raise IndexError(f"文件[{file}]中不存在索引为 [{sheet_index}] 的工作表")
            return workbook, workbook.worksheets[sheet_index]

        # sheet_name和sheet_index都没有传递
        raise ValueError("必须提供sheet_name或sheet_index中的一个")

    @staticmethod
    def get_cell_value(sheet: Worksheet, cell: Cell, default: str = "") -> str:
        """
        获得单元格的值,如果是合并单元格取起始位置的值
        :param sheet: Worksheet
        :param cell: 要获取值的单元格
        :param default: 默认值,缺省为空字符串
        :return: 单元格值
        """
        value = default

        # 单元格有值
        if cell.value is not None:
            value = str(cell.value)
        else:
            # 单元格无值看是否在合并单元格中
            for merged_range in sheet.merged_cells.ranges:
                if (merged_range.min_row <= cell.row <= merged_range.max_row and
                        merged_range.min_col <= cell.column <= merged_range.max_col):
                    value = str(sheet.cell(merged_range.min_row, merged_range.min_col).value)

        # 单元格无值且不在合并单元格中
        return value.strip()

    @staticmethod
    def get_excel_column_values(file: str, sheet_index: int, column_index: int, start_row_index: int = 1) -> List[str]:
        """
        获得指定excel sheet column值集合
        :param file: excel文件
        :param sheet_index: sheet索引
        :param column_index: 列索引
        :param start_row_index: 起始行索引
        :return: 列值集合
        """
        workbook: Workbook
        with UtilsExcel.open_workbook(file, read_only=False) as workbook:
            cells = workbook.worksheets[sheet_index].iter_cols(
                min_col=column_index,
                max_col=column_index,
                min_row=start_row_index,
                values_only=True
            )
            values = list(next(cells))

            # 过滤掉 None 和空字符串，转换为 str，集合{}自动去重
            """
            List 和 list 的区别
            1. list → Python 内置类型，用来创建列表，例如 list("abc") → ['a','b','c']。
            2. List → 来自 typing 模块，只能用来做类型注解，不能当作函数来调用。            
            """
            return list({str(v).strip() for v in values if v not in (None, "")})

    @staticmethod
    def write_excel(model: WriteExcelModel, func: Callable[[Worksheet], None]):
        """
        写excel
        :param model: 写excel文件模型信息
        :param func: 操作sheet的回调函数
        :return: None
        """
        workbook: Workbook
        with UtilsExcel.open_workbook(model.file, read_only=False) as workbook:
            if model.sheet_name:
                if model.sheet_name not in workbook.sheetnames:
                    raise ValueError(f"文件[{model.file}]不包含[{model.sheet_name}]Sheet")
                sheet = workbook[model.sheet_name]
            else:
                if len(workbook.sheetnames) < model.sheet_index - 1:
                    raise IndexError(f"文件[{model.file}]不包含索引位置[{model.sheet_index}]Sheet")
                sheet = workbook.worksheets[model.sheet_index]

            # 回调写入sheet
            func(sheet)
            workbook.save(model.saveas_file if model.saveas_file else model.file)
