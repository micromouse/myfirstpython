"""
使用xlwings控制excel演示
"""
from pathlib import Path

import xlwings
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

class ExcelDemo:
    """
    Excel演示类
    """

    @staticmethod
    def write_excel_by_openpyxl(file: str) -> None:
        """
        使用openpyxl库写excel
        :param file: excel文件名
        :return: None
        """
        workbook: Workbook = load_workbook(file, read_only=False)
        sheet: Worksheet = workbook["Sheet1"]

        # 读合并单元格
        a1 = ExcelDemo._get_cell_value(sheet, 1, 1)
        a2 = ExcelDemo._get_cell_value(sheet, 2, 1)
        a3 = ExcelDemo._get_cell_value(sheet, 3, 1)

        # 添加新行
        sheet.insert_rows(17, 1)

        # 设置单元格格式
        sheet.cell(17, 2, "test")
        sheet["D60"] = "hello"

        # 合并单元格，设置合并单元格值只能是起始单元格
        sheet.merge_cells(range_string="A1:E1")
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
        sheet["A1"] = "cc"
        sheet["A2"] = "cc"

        # 保存
        workbook.save(file)
        print("已保存到文件", file)

    @staticmethod
    def create_excel(file: str):
        """
        建立excel文件
        :param file: 文件名
        """
        # 自动建立目录
        ExcelDemo._create_folder(file)

        # 创建excel文件
        with xlwings.App(visible=False) as app:
            with app.books.add() as book:
                sheet = book.sheets[0]

                # 写入内容
                sheet.range("A1").value = "姓名"
                sheet.range("B1").value = "年龄"
                sheet.range("A2").value = "张三"
                sheet.range("B2").value = 20

                # 保存文件
                book.save(file)

    @staticmethod
    def _get_cell_value(sheet: Worksheet, row: int, column: int) -> str | None:
        """
        获得单元格的值,如果是合并单元格取起始位置的值
        :param sheet: 工作表
        :param row: 行号
        :param column: 列号
        :return: 单元格值
        """
        cell = sheet.cell(row=row, column=column)

        # 单元格有值
        if cell.value is not None:
            return cell.value

        # 单元格无值看是否在合并单元格中
        for merged_range in sheet.merged_cells.ranges:
            if (merged_range.min_row <= row <= merged_range.max_row and
                    merged_range.min_col <= column <= merged_range.max_col):
                return sheet.cell(merged_range.min_row, merged_range.min_col).value

        # 单元格无值且不在合并单元格中
        return None

    @staticmethod
    def _create_folder(file: str):
        path = Path(file)

        # 文件存在抛出文件存在异常
        if path.exists() and path.is_file():
            raise FileExistsError(f"文件[{file}]已存在")

        # 目录不存在自动创建目录
        # parents=True 表示创建所有父目录，exist_ok=True 表示目录已存在也不会报错
        path.parent.mkdir(parents=True, exist_ok=True)

# 测试
ExcelDemo.write_excel_by_openpyxl("D:\\excel.xlsx")
